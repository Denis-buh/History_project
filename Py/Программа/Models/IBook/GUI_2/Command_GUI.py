



from PyQt5.QtWidgets import QFileDialog, QMessageBox
import io
from PIL import Image
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment



class Command_GUI:
    # Добавить обложку книги
    def append_image(self):
        try:
            # Получаем путь к файлу
            fname = QFileDialog.getOpenFileName(self.window, 'Добавить обложку книги', '', 'Расширения изображений (*.png; *.jpg);;Все_файлы (*)')[0]
            # Открываем изображение в PIL
            img = Image.open(fname)
            # изменяем размер изображения
            new_image = img.resize((400, 400))
            byte_io = BytesIO()
            # Сохраняем изображение в ячейку
            new_image.save(byte_io, 'PNG')
            # Читаем изображение в ячейке (бинари формат)
            read = byte_io.getvalue()
            data = io.BytesIO(read).getvalue()

            # Узнаем действующую вкладку
            ind = self.GUI_window.TAB.get_ind()
            vidget = self.sp_widget_book[ind]
            window, file_GUI, name_book, con, cur = vidget
            file_GUI.image_book.load_image(data)
        except:
            msg = QMessageBox()
            # Указываем тип ошибки (Значок). Конструкция: QMessageBox."Искать нужное в документации"
            msg.setIcon(QMessageBox.Warning)
            msg.setText('Невозможно открыть данный файл')
            # Используемые кнопки. Конструкция: QMessageBox."Искать нужное в документации" | QMessageBox."Искать нужное в документации"
            # Для коректного отображения
            msg.exec_()

        
    
    # Показываем должников
    def show_who_take(self):
        for i in self.sp_widget_book:
            if len(i) == 4:
                msg = QMessageBox()
                # Указываем тип ошибки (Значок). Конструкция: QMessageBox."Искать нужное в документации"
                msg.setIcon(QMessageBox.Warning)
                msg.setText("Данная вкладка уже есть")
                # Для коректного отображения
                msg.exec_()
                return 
        # Узнаем действующую вкладку
        ind = self.GUI_window.TAB.get_ind()
        vidget = self.sp_widget_book[ind]
        window, file_GUI, name_book, con, cur = vidget
        self.book_GUI_3.show(name_book, con, cur)
        
    def made_othet(self):
        '''Делаем отчет'''
        # Узнаем действующую вкладку
        ind = self.GUI_window.TAB.get_ind()
        vidget = self.sp_widget_book[ind]
        window, file_GUI, name_book, con, cur = vidget
        
        


        excel_file = Workbook()
        excel_sheet = excel_file.create_sheet(title='Отчет по книге {0}'.format(name_book), index=0)
        x = 0
        y = 1
        sp = ['ФИО', 'Телефон', 'Класс', 'Сколько взято', 'Когда взял', 'Срок (в месяцах)', 'Когда нужно вернуть']

        
        excel_sheet.column_dimensions['A'].width = 50
        excel_sheet.column_dimensions['B'].width = 25
        excel_sheet.column_dimensions['D'].width = 15
        excel_sheet.column_dimensions['E'].width = 25
        excel_sheet.column_dimensions['F'].width = 20
        excel_sheet.column_dimensions['G'].width = 25


        x += 1
        excel_sheet.cell(row=x, column=y).value = 'Должники по книге: {0}'.format(name_book)
        excel_sheet.merge_cells(start_row=x, start_column=y, end_row=x, end_column=y + 6)
        excel_page = excel_sheet.cell(row=x, column=y)
        excel_page.font = Font(bold=True) 
        excel_page.alignment = Alignment(horizontal='center')
        
        x += 1
        ky = y - 1
        for ki in sp:
           ky += 1
           excel_sheet.cell(row=x, column=ky).value = ki 
           
        inform = cur.execute('''SELECT * FROM who_take
                                    WHERE Name_book == "{0}"'''.format(name_book)).fetchall()
        print(inform)
        if len(inform) > 0:
            for i in inform:
                
                x += 1
                excel_sheet.cell(row=x, column=y).value = i[0]
                # excel_sheet.cell(row=x, column=y + 1).value = inform[1]
                excel_sheet.cell(row=x, column=y + 2 - 1).value = i[2]
                excel_sheet.cell(row=x, column=y + 3 - 1).value = i[3]
                excel_sheet.cell(row=x, column=y + 4 - 1).value = i[4]
                excel_sheet.cell(row=x, column=y + 5 - 1).value = i[5]
                excel_sheet.cell(row=x, column=y + 6 - 1).value = i[6]
                excel_sheet.cell(row=x, column=y + 7 - 1).value = i[7]
        x += 1
        excel_sheet.cell(row=x, column=y).value = ""
        

        
        # save the file
        fname = QFileDialog.getSaveFileName(self.window, 'Сохраните отчет', "Отчет по книге {0}.xlsx".format(name_book),
                                            'Расширение_файла (*.xlsx);;Все_файлы (*)')[0]
        if len(fname) == 0:
            return 
        excel_file.save(filename = fname)
        

        

    # Показываем информацию о книге
    def inform_show(self, file_GUI, name_book, type, inform, image, have, how_take):
        file_GUI.le_name_book.setText(str(name_book))
        file_GUI.le_type.setText(str(type))
        file_GUI.pte_inform.insertPlainText(str(inform))
        file_GUI.how_have.setValue(have)
        file_GUI.how_not_see.setText(str(how_take))
        file_GUI.how_see.setText(str(have - how_take))
        
        data = io.BytesIO(image).getvalue()
        
        file_GUI.image_book.load_image(data)
        
    # Возвращаем информацию с вкладки
    def return_inform(self, file_GUI):
        name_book = file_GUI.le_name_book.text()
        
        type_book = file_GUI.le_type.text()
        if type_book == "":
            type_book = None
            
        inform_book = file_GUI.pte_inform.toPlainText()
        if inform_book == "":
            inform_book = "Информация отсутствует"
        how_have = file_GUI.how_have.value()
        image = file_GUI.image_book.get_image()
        return name_book, type_book, inform_book, how_have, image