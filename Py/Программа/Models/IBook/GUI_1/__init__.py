



# Импорт библиотек
from  .IBook_name_GUI import *
from .Command_GUI import Command_Book_GUI_1
from .Command_Table import Command_Table
from PyQt5.QtWidgets import *
from PyQt5.QtSql import *
from PyQt5.QtGui import *
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


class Book_GUI_1(Command_Table, Command_Book_GUI_1):
    def __init__(self, window, GUI_window, sp_widget_book, book_GUI_2):
        # Основное окно 
        self.window = window
        # GUI настройки основного окна
        self.GUI_window = GUI_window
        # Список виджетов для основного окна
        self.sp_widget_book = sp_widget_book
        # Класс GUI окон 2 порядка
        self.book_GUI_2 = book_GUI_2
        

    # Показываем вкладку в окне
    def show(self, con, cur):
        window = QWidget()
        file_GUI = Ui_Form()
        file_GUI.setupUi(window)
        file_GUI.view.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        window.setLayout(file_GUI.main_box)
        self.made_names(cur)
        self.show_inform(file_GUI, self.get_names(cur))
        self.connect(file_GUI)
        self.GUI_window.show_tab(window, "Список книг")
        self.sp_widget_book.append([window, file_GUI, "GUI_1", con, cur])
        
    # Добавляем команды к кнопкам
    def connect(self, file_GUI):
        file_GUI.bt_open.clicked.connect(self.current_item_changed)
        file_GUI.bt_append.clicked.connect(self.append_book)
        file_GUI.bt_dell.clicked.connect(self.del_focus)
        file_GUI.bt_new_inform.clicked.connect(self.show_new_inform)
        file_GUI.bt_found.clicked.connect(self.found_book)
        file_GUI.bt_oth.clicked.connect(self.made_othet)
        
    def made_othet(self):
        '''Делаем отчет'''
        # Узнаем действующую вкладку
        ind = self.GUI_window.TAB.get_ind()
        vidget = self.sp_widget_book[ind]
        window, file_GUI, kesh, con, cur = vidget
        Name_books = cur.execute('''SELECT Name_book FROM inform''').fetchall()
        print(Name_books)

        excel_file = Workbook()
        excel_sheet = excel_file.create_sheet(title='Отчет по книгам', index=0)
        x = 0
        y = 1
        sp = ['ФИО', 'Телефон', 'Класс', 'Сколько взято', 'Когда взял', 'Срок (в месяцах)', 'Когда нужно вернуть']
        x += 1
        
        excel_sheet.column_dimensions['A'].width = 50
        excel_sheet.column_dimensions['B'].width = 25
        excel_sheet.column_dimensions['D'].width = 15
        excel_sheet.column_dimensions['E'].width = 25
        excel_sheet.column_dimensions['F'].width = 20
        excel_sheet.column_dimensions['G'].width = 25
        
        
        excel_sheet.cell(row=x, column=y).value = 'Отчет по книгам'
        excel_page = excel_sheet.cell(row=x, column=y)
        excel_page.font = Font(size=14, bold=True) 
        excel_page.alignment = Alignment(horizontal='center')
        excel_sheet.merge_cells(start_row=x, start_column=y, end_row=x, end_column=y + 6)
        
        x += 1
        excel_sheet.cell(row=x, column=y).value = ""
        
        for i in Name_books:
            x += 1
            excel_sheet.cell(row=x, column=y).value = 'Должники по книге: {0}'.format(i[0])
            excel_sheet.merge_cells(start_row=x, start_column=y, end_row=x, end_column=y + 6)
            excel_page = excel_sheet.cell(row=x, column=y)
            excel_page.font = Font(bold=True) 
            excel_page.alignment = Alignment(horizontal='center')
            
            x += 1
            ky = y - 1
            for ki in sp:
               ky += 1
               excel_sheet.cell(row=x, column=ky).value = ki 
               
            inform = cur.execute('''SELECT * FROM who_take
                                        WHERE Name_book == "{0}"'''.format(i[0])).fetchall()
            print(inform)
            if len(inform) > 0:
                for i in inform:
                    x += 1
                    excel_sheet.cell(row=x, column=y).value = i[0]
                    # excel_sheet.cell(row=x, column=y + 1).value = inform[1]
                    excel_sheet.cell(row=x, column=y + 2 - 1).value = i[2]
                    excel_sheet.cell(row=x, column=y + 3 - 1).value = i[3]
                    excel_sheet.cell(row=x, column=y + 4 - 1).value = i[4]
                    excel_sheet.cell(row=x, column=y + 5 - 1).value = i[5]
                    excel_sheet.cell(row=x, column=y + 6 - 1).value = i[6]
                    excel_sheet.cell(row=x, column=y + 7 - 1).value = i[7]
            x += 1
            excel_sheet.cell(row=x, column=y).value = ""
        

        
        # save the file
        fname = QFileDialog.getSaveFileName(self.window, 'Сохраните отчет', "Отчет по книгам.xlsx",
                                            'Расширение_файла (*.xlsx);;Все_файлы (*)')[0]
        if len(fname) == 0:
            return 
        excel_file.save(filename = fname)
